import os
import json
import traceback
import boto3
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ─── 환경변수 ───
S3_BUCKET = os.environ.get('S3_BUCKET', '')

# ─── 색상 상수 ───
BLUE_DARK    = "1E3A5F"
BLUE_MID     = "2563EB"
BLUE_LIGHT   = "DBEAFE"
KAKAO_DARK   = "92400E"
KAKAO_LIGHT  = "FEF3C7"
KAKAO_YELLOW = "F59E0B"
GRAY_BG      = "F8FAFC"
GRAY_BORDER  = "CBD5E1"
WHITE        = "FFFFFF"
GREEN        = "059669"
RED          = "DC2626"

# ─── AWS 클라이언트 ───
s3 = boto3.client('s3', region_name='ap-northeast-2')


# ───────────────────────────────
# 스타일 헬퍼
# ───────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _border() -> Border:
    side = Side(style="thin", color=GRAY_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def apply_header(cell, bg=BLUE_DARK):
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=True, color=WHITE)
    cell.alignment = _center()
    cell.border    = _border()


def apply_data(cell, row_idx: int, font_color="000000"):
    cell.fill      = _fill(WHITE if row_idx % 2 == 0 else GRAY_BG)
    cell.font      = _font(color=font_color)
    cell.alignment = _center()
    cell.border    = _border()


def col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def hide_grid(ws):
    ws.sheet_view.showGridLines = False


# ───────────────────────────────
# S3 유틸
# ───────────────────────────────

def load_report_from_s3(conversation_id: str) -> dict:
    key = f"reports/{conversation_id}.json"
    try:
        resp = s3.get_object(Bucket=S3_BUCKET, Key=key)
        return json.loads(resp['Body'].read().decode('utf-8'))
    except s3.exceptions.NoSuchKey:
        raise FileNotFoundError(key)


def upload_xlsx(local_path: str, s3_key: str):
    s3.upload_file(local_path, S3_BUCKET, s3_key)


def presigned_url(s3_key: str, expiry=3600) -> str:
    return s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': S3_BUCKET, 'Key': s3_key},
        ExpiresIn=expiry,
    )


# ───────────────────────────────
# 데이터 파싱
# ───────────────────────────────

def sf(val, default=0.0) -> float:
    """safe float 변환"""
    try:
        return float(val) if val not in (None, '', '-') else default
    except (ValueError, TypeError):
        return default


def extract_data(messages: list) -> dict:
    """messages → {pie, line, bar, table, texts}  (chartType 기반 분류)"""
    result = {
        'pie':   [],
        'line':  [],
        'bar':   [],
        'table': [],
        'texts': [],
    }
    for msg in messages:
        chart_type = msg.get('chartType')
        structured = msg.get('structuredData') or []
        role       = msg.get('role', '')
        content    = msg.get('content', '')

        if chart_type and structured:
            if chart_type in result:
                result[chart_type] = structured
        elif role == 'assistant' and not chart_type and content:
            result['texts'].append(content)

    return result


# ───────────────────────────────
# 시트 1: 📊 요약
# ───────────────────────────────

def build_summary(ws, report: dict, data: dict):
    hide_grid(ws)
    title      = report.get('title', '주간 광고 성과 리포트')
    created    = report.get('createdAt', datetime.now().isoformat())
    pie_rows   = data['pie']
    table_rows = data['table']

    # ── Row 1: 타이틀 ──
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = title
    c.fill  = _fill(BLUE_DARK)
    c.font  = Font(name="Arial", bold=True, color=WHITE, size=18)
    c.alignment = _center()
    ws.row_dimensions[1].height = 36

    # ── Row 2: 생성일 ──
    ws.merge_cells('A2:L2')
    c2 = ws['A2']
    try:
        created_str = datetime.fromisoformat(created).strftime('%Y년 %m월 %d일')
    except Exception:
        created_str = str(created)[:10]
    c2.value = f"생성일: {created_str}"
    c2.fill  = _fill(BLUE_LIGHT)
    c2.font  = _font(color=BLUE_DARK)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    # ── Row 3: 여백 ──
    ws.row_dimensions[3].height = 10

    # ── Row 4: KPI 섹션 레이블 ──
    ws['A4'].value = "▶ 핵심 KPI"
    ws['A4'].font  = _font(bold=True, color=BLUE_DARK, size=12)

    # ── Row 5: KPI 헤더 ──
    for col, h in enumerate(["지표", "이번 주", "전주", "증감", "증감률"], start=1):
        apply_header(ws.cell(row=5, column=col, value=h))

    # ── KPI 수치 계산 ──
    def sum_t(key):
        return sum(sf(r.get(key, 0)) for r in table_rows)

    # pie 데이터에서 매체별 광고비 추출
    g_pie  = next((r for r in pie_rows if r.get('매체') == '구글'),   {})
    k_pie  = next((r for r in pie_rows if r.get('매체') == '카카오'), {})
    g_cost = sf(g_pie.get('광고비(원)', 0))
    k_cost = sf(k_pie.get('광고비(원)', 0))

    # pie 데이터 없으면 table에서 광고비 보완
    if g_cost == 0 and k_cost == 0:
        g_cost = sum_t('구글 광고비(원)')
        k_cost = sum_t('카카오 광고비(원)')

    # table 데이터에서 노출/클릭 합산
    g_imp   = sum_t('구글 노출')
    g_click = sum_t('구글 클릭')
    k_imp   = sum_t('카카오 노출')
    k_click = sum_t('카카오 클릭')

    total_cost  = g_cost + k_cost
    total_imp   = g_imp + k_imp
    total_click = g_click + k_click
    avg_ctr     = total_click / total_imp if total_imp > 0 else 0

    kpi_rows = [
        ("총 광고비(원)",     total_cost,  "#,##0"),
        ("총 노출수",          total_imp,   "#,##0"),
        ("총 클릭수",          total_click, "#,##0"),
        ("평균 CTR",           avg_ctr,     "0.00%"),
        ("구글 광고비(원)",    g_cost,      "#,##0"),
        ("카카오 광고비(원)",  k_cost,      "#,##0"),
    ]

    for idx, (label, value, fmt) in enumerate(kpi_rows):
        r = 6 + idx
        c_label = ws.cell(row=r, column=1, value=label)
        apply_data(c_label, idx)

        c_val = ws.cell(row=r, column=2, value=round(value, 4))
        apply_data(c_val, idx)
        c_val.number_format = fmt

        for col in range(3, 6):
            c = ws.cell(row=r, column=col, value="-")
            apply_data(c, idx)

    # ── Row 12: 여백 ──
    ws.row_dimensions[12].height = 10

    # ── Row 13: 매체별 섹션 레이블 ──
    ws['A13'].value = "▶ 매체별 광고비 비중"
    ws['A13'].font  = _font(bold=True, color=BLUE_DARK, size=12)

    # ── Row 14: 헤더 ──
    for col, h in enumerate(["매체", "이번주 광고비(원)", "비중(%)"], start=1):
        apply_header(ws.cell(row=14, column=col, value=h))

    # ── Row 15~16: 구글/카카오 ──
    for idx, (media, cost) in enumerate([("구글", round(g_cost, 0)), ("카카오", round(k_cost, 0))]):
        r = 15 + idx
        c1 = ws.cell(row=r, column=1, value=media)
        apply_data(c1, idx)
        c2 = ws.cell(row=r, column=2, value=cost)
        c2.number_format = "#,##0"
        apply_data(c2, idx)
        c3 = ws.cell(row=r, column=3, value=f"=B{r}/SUM($B$15:$B$16)")
        c3.number_format = "0.00%"
        apply_data(c3, idx)

    # ── 파이 차트: E13 ──
    pie = PieChart()
    pie.title  = "매체별 광고비 비중"
    pie.style  = 10
    pie_data   = Reference(ws, min_col=2, min_row=14, max_row=16)
    labels     = Reference(ws, min_col=1, min_row=15, max_row=16)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(labels)
    pt_g = DataPoint(idx=0)
    pt_g.graphicalProperties.solidFill = BLUE_MID
    pt_k = DataPoint(idx=1)
    pt_k.graphicalProperties.solidFill = KAKAO_YELLOW
    pie.series[0].dPt = [pt_g, pt_k]
    pie.width  = 12
    pie.height = 10
    ws.add_chart(pie, "E13")

    # ── Row 22: 주차별 섹션 레이블 ──
    ws['A22'].value = "▶ 주차별 성과 비교"
    ws['A22'].font  = _font(bold=True, color=BLUE_DARK, size=12)

    # ── Row 23: 헤더 ──
    weekly_headers = ["주차", "구글 광고비(원)", "카카오 광고비(원)", "총 광고비(원)",
                       "구글 클릭", "카카오 클릭", "총 클릭", "비중(%)"]
    for col, h in enumerate(weekly_headers, start=1):
        apply_header(ws.cell(row=23, column=col, value=h))

    # ── Row 24~28: 주차별 데이터 (table 데이터에서 날짜 → 주차 집계) ──
    weekly_map: dict = {}
    for row in table_rows:
        date_str = row.get('날짜', '')
        try:
            dt       = datetime.strptime(date_str, '%Y-%m-%d')
            week_num = (dt.day - 1) // 7 + 1
            wlabel   = f"{dt.month}월 {week_num}주차"
        except Exception:
            wlabel = date_str or '기타'
        entry = weekly_map.setdefault(wlabel, {'g_cost': 0, 'k_cost': 0, 'g_click': 0, 'k_click': 0})
        entry['g_cost']  += sf(row.get('구글 광고비(원)', 0))
        entry['k_cost']  += sf(row.get('카카오 광고비(원)', 0))
        entry['g_click'] += sf(row.get('구글 클릭', 0))
        entry['k_click'] += sf(row.get('카카오 클릭', 0))

    weekly_items = list(weekly_map.items())[:5]
    if not weekly_items:
        weekly_items = [(f"W{i}", {'g_cost': 0, 'k_cost': 0, 'g_click': 0, 'k_click': 0}) for i in range(1, 6)]

    for idx, (wlabel, entry) in enumerate(weekly_items):
        r    = 24 + idx
        g_c  = entry['g_cost']
        k_c  = entry['k_cost']
        g_cl = entry['g_click']
        k_cl = entry['k_click']

        row_vals = [wlabel, g_c, k_c, f"=B{r}+C{r}", g_cl, k_cl, f"=E{r}+F{r}",
                    f"=D{r}/SUM($D$24:$D$28)"]
        for col, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            apply_data(c, idx)
            if col in (2, 3, 4):
                c.number_format = "#,##0"
            elif col == 8:
                c.number_format = "0.00%"

    # ── 막대 차트: I22 ──
    bar = BarChart()
    bar.type      = "col"
    bar.grouping  = "clustered"
    bar.title     = "주차별 광고비 비교"
    bar.style     = 10
    bar.y_axis.title = "광고비(원)"
    cats  = Reference(ws, min_col=1, min_row=24, max_row=28)
    g_ref = Reference(ws, min_col=2, min_row=23, max_row=28)
    k_ref = Reference(ws, min_col=3, min_row=23, max_row=28)
    bar.add_data(g_ref, titles_from_data=True)
    bar.add_data(k_ref, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = BLUE_MID
    bar.series[1].graphicalProperties.solidFill = KAKAO_YELLOW
    bar.width  = 16
    bar.height = 10
    ws.add_chart(bar, "I22")

    col_widths(ws, [18, 16, 16, 16, 16, 5, 5, 5, 5, 5, 5, 5])


# ───────────────────────────────
# 시트 2: 📈 일별 추이
# ───────────────────────────────

def _find_date(row: dict) -> str:
    for v in row.values():
        if isinstance(v, str) and DATE_PAT.match(v):
            return v
    return row.get('date', row.get('basic_date', 'unknown'))


def build_daily(ws, report: dict, data: dict):
    hide_grid(ws)
    line_rows  = data['line']
    table_rows = data['table']

    # ── Row 1: 타이틀 ──
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = "일별 광고 성과 추이"
    c.fill  = _fill(BLUE_DARK)
    c.font  = Font(name="Arial", bold=True, color=WHITE, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws['A2'].value = "■ 이번 주"
    ws['A2'].font  = _font(bold=True, color=BLUE_DARK)

    # ── Row 3: 컬럼 헤더 ──
    daily_hdrs = ["날짜", "구글 광고비(원)", "카카오 광고비(원)", "총 광고비(원)",
                  "구글 노출", "구글 클릭", "카카오 노출", "카카오 클릭",
                  "총 노출(합계)", "총 클릭(합계)"]
    for col, h in enumerate(daily_hdrs, start=1):
        apply_header(ws.cell(row=3, column=col, value=h))

    # ── 날짜별 집계 ──
    # line 데이터: 구글/카카오 광고비
    date_map: dict[str, dict] = {}
    for row in line_rows:
        d = row.get('날짜', '')
        if not d:
            continue
        date_map.setdefault(d, {'g_cost': 0, 'k_cost': 0, 'g_imp': 0, 'g_click': 0, 'k_imp': 0, 'k_click': 0})
        date_map[d]['g_cost'] += sf(row.get('구글 광고비(원)', 0))
        date_map[d]['k_cost'] += sf(row.get('카카오 광고비(원)', 0))

    # table 데이터: 노출/클릭 보완
    table_by_date = {r.get('날짜', ''): r for r in table_rows if r.get('날짜')}
    for d, m in date_map.items():
        t = table_by_date.get(d, {})
        m['g_imp']   = sf(t.get('구글 노출', 0))
        m['g_click'] = sf(t.get('구글 클릭', 0))
        m['k_imp']   = sf(t.get('카카오 노출', 0))
        m['k_click'] = sf(t.get('카카오 클릭', 0))

    sorted_dates = sorted(date_map.keys())
    DATA_START   = 4

    if not sorted_dates:
        ws.cell(row=DATA_START, column=1, value="데이터 없음")
        data_end = DATA_START
    else:
        for idx, date_val in enumerate(sorted_dates):
            r = DATA_START + idx
            m = date_map[date_val]
            vals = [date_val,
                    round(m['g_cost'], 0), round(m['k_cost'], 0), f"=B{r}+C{r}",
                    round(m['g_imp'], 0),  round(m['g_click'], 0),
                    round(m['k_imp'], 0),  round(m['k_click'], 0),
                    f"=E{r}+G{r}", f"=F{r}+H{r}"]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=r, column=col, value=val)
                apply_data(c, idx)
                if col in (2, 3, 4):
                    c.number_format = "#,##0"

        data_end = DATA_START + len(sorted_dates) - 1
        sum_r    = data_end + 1

        # 합계 행
        ws.cell(row=sum_r, column=1, value="합계")
        for col in range(1, 11):
            c = ws.cell(row=sum_r, column=col)
            c.fill = _fill(BLUE_LIGHT)
            c.font = _font(bold=True, color=BLUE_DARK)
            c.border = _border()
            c.alignment = _center()
        for col in range(2, 11):
            col_l = get_column_letter(col)
            c = ws.cell(row=sum_r, column=col,
                        value=f"=SUM({col_l}{DATA_START}:{col_l}{data_end})")
            if col in (2, 3, 4):
                c.number_format = "#,##0"

        # 자동필터
        ws.auto_filter.ref = f"A3:J{data_end}"

        # ── 라인 차트: L4 ──
        line = LineChart()
        line.title        = "총 광고비 추이 (이번 주)"
        line.style        = 10
        line.y_axis.title = "광고비(원)"
        line.x_axis.title = "날짜"
        cost_ref = Reference(ws, min_col=4, min_row=3, max_row=data_end)
        line.add_data(cost_ref, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=DATA_START, max_row=data_end)
        line.set_categories(cats)
        line.series[0].graphicalProperties.line.solidFill = BLUE_MID
        line.width  = 24
        line.height = 14
        ws.add_chart(line, "L4")
        data_end = sum_r

    # ── 전주 섹션 ──
    prev_row = data_end + 4
    ws.cell(row=prev_row, column=1, value="■ 전주").font = _font(bold=True, color=BLUE_DARK)
    ws.cell(row=prev_row + 1, column=1,
            value="전주 데이터 없음 (시스템에 누적되면 자동 표시됩니다)").font = \
        _font(color="666666", italic=True)

    col_widths(ws, [14, 16, 16, 16, 12, 12, 12, 12, 14, 14])


# ───────────────────────────────
# 시트 3: 🔵 구글 상세
# ───────────────────────────────

def _formula_cell(ws, row, col, formula, fmt, idx):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = fmt
    c.fill      = _fill(WHITE if idx % 2 == 0 else GRAY_BG)
    c.font      = _font()
    c.border    = _border()
    c.alignment = _center()


def build_google(ws, data: dict):
    hide_grid(ws)
    g_rows = data['bar']

    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = "Google Ads 캠페인 상세 성과"
    c.fill  = _fill(BLUE_DARK)
    c.font  = Font(name="Arial", bold=True, color=WHITE, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    g_hdrs = ["캠페인명", "노출수", "클릭수", "CTR", "광고비(원)", "CPC", "전환수", "전환가치(원)", "ROAS"]
    for col, h in enumerate(g_hdrs, start=1):
        apply_header(ws.cell(row=2, column=col, value=h))

    if not g_rows:
        ws.cell(row=3, column=1, value="데이터 없음")
        return

    DATA_START = 3
    for idx, row in enumerate(g_rows):
        r        = DATA_START + idx
        imp      = sf(row.get('노출수', 0))
        click    = sf(row.get('클릭수', 0))
        cost     = sf(row.get('광고비(원)', 0))
        conv     = sf(row.get('전환수', 0))
        conv_val = sf(row.get('전환가치(원)', 0))

        plain = {1: row.get('캠페인명', '(미분류)'), 2: imp, 3: click, 5: round(cost, 0), 7: conv, 8: round(conv_val, 0)}
        for col, val in plain.items():
            c = ws.cell(row=r, column=col, value=val)
            apply_data(c, idx)

        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"
        ws.cell(row=r, column=7).number_format = "#,##0"
        ws.cell(row=r, column=8).number_format = "#,##0"

        _formula_cell(ws, r, 4, f"=IF(B{r}=0,0,C{r}/B{r})",          "0.00%",  idx)
        _formula_cell(ws, r, 6, f'=IF(C{r}=0,"-",E{r}/C{r})',         "#,##0",  idx)
        _formula_cell(ws, r, 9, f'=IF(E{r}=0,"-",H{r}/E{r})',         "0.0%",   idx)

    data_end = DATA_START + len(g_rows) - 1
    sum_r    = data_end + 1

    # 합계 행
    ws.cell(row=sum_r, column=1, value="합계")
    for col in range(1, 10):
        c = ws.cell(row=sum_r, column=col)
        c.fill = _fill(BLUE_LIGHT)
        c.font = _font(bold=True, color=BLUE_DARK)
        c.border = _border()
        c.alignment = _center()
    for col in [2, 3, 5, 7, 8]:
        col_l = get_column_letter(col)
        c = ws.cell(row=sum_r, column=col,
                    value=f"=SUM({col_l}{DATA_START}:{col_l}{data_end})")
        c.number_format = "#,##0"

    ws.auto_filter.ref = f"A2:I{data_end}"

    # ── 가로 막대 차트 ──
    bar = BarChart()
    bar.type      = "bar"
    bar.grouping  = "clustered"
    bar.title     = "캠페인별 클릭수"
    bar.style     = 10
    click_ref = Reference(ws, min_col=3, min_row=2, max_row=data_end)
    bar.add_data(click_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=DATA_START, max_row=data_end)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = BLUE_MID
    bar.width  = 18
    bar.height = 12
    ws.add_chart(bar, f"A{sum_r + 2}")

    col_widths(ws, [24, 12, 12, 10, 14, 12, 10, 14, 10])


# ───────────────────────────────
# 시트 4: 🟡 카카오 상세
# ───────────────────────────────

def build_kakao(ws, data: dict):
    hide_grid(ws)
    pie_rows = data['pie']

    # ── Row 1: 타이틀 ──
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = "카카오 광고 캠페인 상세 성과"
    c.fill  = _fill(KAKAO_DARK)
    c.font  = Font(name="Arial", bold=True, color=WHITE, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    k_hdrs = ["캠페인명", "노출수", "클릭수", "CTR", "광고비(원)", "CPC", "CPM", "도달(Reach)"]
    for col, h in enumerate(k_hdrs, start=1):
        apply_header(ws.cell(row=2, column=col, value=h), bg=KAKAO_DARK)

    # ── Row 3: 데이터 없음 안내 ──
    notice = ws.cell(row=3, column=1,
                     value="카카오 캠페인 상세 데이터가 없습니다. (캠페인별 분리 데이터가 제공되면 자동으로 표시됩니다)")
    notice.font      = _font(color=KAKAO_DARK, italic=True)
    notice.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells('A3:H3')
    ws.row_dimensions[3].height = 22

    # ── Row 5: 매체 합계 섹션 (pie 데이터 활용) ──
    ws['A5'].value = "▶ 카카오 광고비 합계 (매체별 요약)"
    ws['A5'].font  = _font(bold=True, color=KAKAO_DARK, size=12)

    for col, h in enumerate(["매체", "광고비(원)"], start=1):
        apply_header(ws.cell(row=6, column=col, value=h), bg=KAKAO_DARK)

    k_pie = next((r for r in pie_rows if r.get('매체') == '카카오'), {})
    k_total = sf(k_pie.get('광고비(원)', 0))

    c1 = ws.cell(row=7, column=1, value="카카오")
    apply_data(c1, 0)
    c2 = ws.cell(row=7, column=2, value=round(k_total, 0))
    c2.number_format = "#,##0"
    apply_data(c2, 0)

    # ── 파이 차트: 매체별 광고비 비중 (pie 데이터 전체) ──
    if len(pie_rows) >= 2:
        # 요약 시트의 B15:B16을 직접 참조하지 않고 현 시트에 데이터 작성 후 차트 연결
        ws['A9'].value  = "▶ 매체별 광고비 비중 (참고)"
        ws['A9'].font   = _font(bold=True, color=KAKAO_DARK, size=12)
        for col, h in enumerate(["매체", "광고비(원)"], start=1):
            apply_header(ws.cell(row=10, column=col, value=h), bg=KAKAO_DARK)

        for idx, row in enumerate(pie_rows):
            r  = 11 + idx
            c1 = ws.cell(row=r, column=1, value=row.get('매체', f'매체{idx+1}'))
            apply_data(c1, idx)
            c2 = ws.cell(row=r, column=2, value=round(sf(row.get('광고비(원)', 0)), 0))
            c2.number_format = "#,##0"
            apply_data(c2, idx)

        pie_end = 10 + len(pie_rows)
        pie_chart = PieChart()
        pie_chart.title  = "매체별 광고비 비중"
        pie_chart.style  = 10
        pie_data  = Reference(ws, min_col=2, min_row=10, max_row=pie_end)
        labels    = Reference(ws, min_col=1, min_row=11, max_row=pie_end)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pt_g = DataPoint(idx=0)
        pt_g.graphicalProperties.solidFill = BLUE_MID
        pt_k = DataPoint(idx=1)
        pt_k.graphicalProperties.solidFill = KAKAO_YELLOW
        pie_chart.series[0].dPt = [pt_g, pt_k]
        pie_chart.width  = 16
        pie_chart.height = 12
        ws.add_chart(pie_chart, "D9")

    col_widths(ws, [40, 16, 12, 10, 14, 12, 12, 12])


# ───────────────────────────────
# Lambda Handler
# ───────────────────────────────

def lambda_handler(event, context):
    print(f"Excel Export Lambda 시작: {json.dumps(event, ensure_ascii=False)}")

    conversation_id = event.get('conversationId', '').strip()
    if not conversation_id:
        return {
            'statusCode': 400,
            'body': json.dumps(
                {'error': 'INVALID_INPUT', 'message': 'conversationId가 필요합니다'},
                ensure_ascii=False,
            ),
        }

    # ── 1. S3에서 리포트 JSON 로드 ──
    try:
        report = load_report_from_s3(conversation_id)
    except FileNotFoundError:
        return {
            'statusCode': 404,
            'body': json.dumps(
                {'error': 'REPORT_NOT_FOUND', 'message': '리포트 데이터를 찾을 수 없습니다'},
                ensure_ascii=False,
            ),
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps({'error': 'PARSE_ERROR', 'message': str(e)}, ensure_ascii=False),
        }

    try:
        # ── 2. 데이터 파싱 ──
        messages = report.get('messages', [])
        data     = extract_data(messages)
        title    = report.get('title', '주간리포트')

        # ── 3. Excel 워크북 생성 ──
        wb  = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "📊 요약"
        ws2 = wb.create_sheet("📈 일별 추이")
        ws3 = wb.create_sheet("🔵 구글 상세")
        ws4 = wb.create_sheet("🟡 카카오 상세")

        build_summary(ws1, report, data)
        build_daily(ws2, report, data)
        build_google(ws3, data)
        build_kakao(ws4, data)

        # ── 4. /tmp 저장 ──
        local_path = f"/tmp/report_{conversation_id}.xlsx"
        wb.save(local_path)
        print(f"Excel 저장: {local_path}")

        # ── 5. S3 업로드 ──
        s3_key = f"excel/report_{conversation_id}.xlsx"
        upload_xlsx(local_path, s3_key)
        print(f"S3 업로드: s3://{S3_BUCKET}/{s3_key}")

        # ── 6. presigned URL ──
        download_url = presigned_url(s3_key)

        # ── 7. 반환 ──
        today_str = datetime.now().strftime('%Y%m%d')
        file_name = f"주간리포트_{title}_{today_str}.xlsx"

        return {
            'statusCode': 200,
            'body': json.dumps(
                {'downloadUrl': download_url, 'fileName': file_name},
                ensure_ascii=False,
            ),
        }

    except Exception as e:
        print(f"Excel 생성 실패: {e}")
        traceback.print_exc()
        return {
            'statusCode': 500,
            'body': json.dumps({'error': 'PARSE_ERROR', 'message': str(e)}, ensure_ascii=False),
        }
